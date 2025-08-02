import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import Data_Tables
import Annual_TIF_Report
from pathlib import Path
import json
import threading, os
import win32com.client as win32

atr_thread: threading.Thread | None = None


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#                                                                       #
#  This section is designed to transfer TIF data from the spreadsheet   #
#  to a list that can be used by the GUI. The list consists of one      #
#  node for each TIF district, with each node consisting of the         #
#  number, name, and group of the district, respectively                #
#                                                                       #
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

class TIF:    
    def __init__(self, data1=None, data2=None, data3=None):
        self.num = data1
        self.name = data2
        self.group = data3
        self.check = 0
        
class TIF_List:
    def __init__(self, file_path=None):
        self.list = []
        self.checkboxes = []
        self.last_clicked = None
        
        if file_path:
            TIF_file = load_workbook(file_path) 
            sheets = TIF_file.sheetnames
            sheet = None
            for cur_sheet in sheets:
                if cur_sheet == 'Section 1':
                    sheet = TIF_file[cur_sheet]
                        
            if sheet is not None:
                row = 2
                while True:
                    num = sheet.cell(row=row, column=1).value
                    if num is None:
                        break
                    name = sheet.cell(row=row, column=2).value
                    group = sheet.cell(row=row, column=7).value
                    
                    self.list.append(TIF(num, name, group))
                    self.checkboxes.append(tk.BooleanVar(value=False))
                    row += 1
        
    def size(self):
        return len(self.list)
    
    def sort_by_column(self, column):
        # Sort the list by the specified column (1: num, 2: name, 3: group)
        if column == 1:
            self.list.sort(key=lambda x: x.num)
        elif column == 2:
            self.list.sort(key=lambda x: x.name)
        elif column == 3:
            self.list.sort(key=lambda x: x.group)
        

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#                                                                       #
#  This section is designed to create the GUI and load in the TIF       #
#  districts so that the program can prepare to run and sort the data   #
#                                                                       #
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

STATE_FILE = "gui_state.json"

def save_state():
    try:
        state = {
            "input_file": input_file if 'input_file' in globals() else "",
            "template_file": template_file if 'template_file' in globals() else "",
            "attB_file": attB_file if 'attB_file' in globals() else "",
            "attB2_file": attB2_file if 'attB2_file' in globals() else "",
            "attC_file": attC_file if 'attC_file' in globals() else "",
            "attC2_file": attC2_file if 'attC2_file' in globals() else "",
            "bsigned_file": bsigned_file if 'bsigned_file' in globals() else "",
            "csigned_file": csigned_file if 'csigned_file' in globals() else "",
            "reporting_year": date_entry.get(),
            "output_data_tables": dt_ceckbox_var.get(),
            "pdf_merger": pdfm_checkbox_var.get()
        }
        with open(STATE_FILE, "w") as f:
            json.dump(state, f)
    except Exception as e:
        print(f"Failed to save state: {e}")

def load_state():
    global input_file, template_file, tif_inst, attB_file, attB2_file, attC_file, attC2_file
    global bsigned_file, csigned_file
    
    if not Path(STATE_FILE).exists():
        return
    try:
        with open(STATE_FILE, "r") as f:
            state = json.load(f)
        input_file = state.get("input_file", "")
        template_file = state.get("template_file", "")
        date_entry.insert(0, state.get("reporting_year", ""))
        dt_ceckbox_var.set(state.get("output_data_tables", False))
        pdfm_checkbox_var.set(state.get("pdf_merger", False))
        
        attB_file    = state.get("attB_file", "")
        attB2_file   = state.get("attB2_file", "")
        attC_file    = state.get("attC_file", "")
        attC2_file   = state.get("attC2_file", "")
        bsigned_file = state.get("bsigned_file", "")
        csigned_file = state.get("csigned_file", "")

        if input_file:
            truncate_label(file_label, text=Path(input_file).name)
            tif_inst = TIF_List(input_file)
            update_grid()

        if template_file:
            truncate_label(template_label, text=Path(template_file).name)
        if bsigned_file:
            truncate_label(bsigned_label, text=Path(bsigned_file).name)
        if csigned_file:
            truncate_label(csigned_label, text=Path(csigned_file).name)
            
        if attB_file:
            truncate_label(attB_label, text=Path(attB_file).name, max_length=12)
        if attB2_file:
            truncate_label(attB2_label, text=Path(attB2_file).name, max_length=12)
        if attC_file:
            truncate_label(attC_label, text=Path(attC_file).name, max_length=12)
        if attC2_file:
            truncate_label(attC2_label, text=Path(attC2_file).name, max_length=12)

    except Exception as e:
        print(f"Failed to load state: {e}")

def truncate_label(label, text, max_length=22):
    if len(text) > max_length:
        text = text[:max_length] + "..."
    label.config(text=text)

def open_input_file():
    global tif_inst
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:  # If a file is selected
        file_name = Path(file_path).name   
        truncate_label(file_label, text=f"{file_name}")
        global input_file
        input_file = file_path
        tif_inst = TIF_List(input_file)
        update_grid()
        
def open_template_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xls *.xlsx *.xlsm"), ("All files", "*.*")]
    )
    if file_path:  # If a file is selected
        file_name = Path(file_path).name
        truncate_label(template_label, text=f"{file_name}")
        global template_file
        template_file = file_path
        
def open_attB_file():
    file_path = filedialog.askopenfilename(filetypes=[("Word documents", "*.docx;*.doc")])
    if file_path:
        file_name = Path(file_path).name
        truncate_label(attB_label, text=f"{file_name}", max_length=12)
        global attB_file
        attB_file = file_path
        
def open_attB2_file():
    file_path = filedialog.askopenfilename(filetypes=[("Word documents", "*.docx;*.doc")])
    if file_path:
        file_name = Path(file_path).name
        truncate_label(attB2_label, text=f"{file_name}", max_length=12)
        global attB2_file
        attB2_file = file_path

def open_attC_file():
    file_path = filedialog.askopenfilename(filetypes=[("Word documents", "*.docx;*.doc")])
    if file_path:
        file_name = Path(file_path).name
        truncate_label(attC_label, text=f"{file_name}", max_length=12)
        global attC_file
        attC_file = file_path
        
def open_attC2_file():
    file_path = filedialog.askopenfilename(filetypes=[("Word documents", "*.docx;*.doc")])
    if file_path:
        file_name = Path(file_path).name
        truncate_label(attC2_label, text=f"{file_name}", max_length=12)
        global attC2_file
        attC2_file = file_path
        
def open_bsigned_file():
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if file_path:  # If a file is selected
        file_name = Path(file_path).name
        truncate_label(bsigned_label, text=f"{file_name}")
        global bsigned_file
        bsigned_file = file_path

def open_csigned_file():
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if file_path:  # If a file is selected
        file_name = Path(file_path).name
        truncate_label(csigned_label, text=f"{file_name}")
        global csigned_file
        csigned_file = file_path
        

word_app: any = None
excel_app: any = None

def cancel_run():
    for app in (word_app, excel_app):
        try:
            if app is not None:
                app.Quit()
        except Exception:
            pass

    # 2) Fallback: force‐kill any leftover processes
    # (Windows only)
    os.system('taskkill /f /im WINWORD.EXE  >nul 2>nul')
    os.system('taskkill /f /im EXCEL.EXE >nul 2>nul')
    
    os._exit(1)        

def run_code():
    global atr_thread, word_app, excel_app
    # Disable the Run button, enable Cancel
    button_run.config(state=tk.DISABLED)
    button_cancel.config(state=tk.NORMAL)
    run_status.config(text="Running...")
    run_status.update_idletasks()
    
    tif_final = []
    for i in range(tif_inst.size()):
        if tif_inst.checkboxes[i].get() == True:
            # print("checked box found")
            tif_final.append(tif_inst.list[i])
    
    if dt_ceckbox_var.get() == True:
        Data_Tables.Data_Tables(date_entry.get())
    
    def target():
        global word_app, excel_app
        # In ATR you do:
        word_app = win32.Dispatch("Word.Application")
        excel_app = win32.Dispatch("Excel.Application")
        
        try:
            Annual_TIF_Report.ATR(tif_final, tif_inst.list, date_entry.get(), input_file, template_file,
                                attB_file, attB2_file, attC_file, attC2_file, bsigned_file, csigned_file,
                                pdfm_checkbox_var.get())
            run_status.config(text="Run Complete")
        except Exception as e:
            run_status.config(text=f"Error: {e}")      
            print(e) 
        finally:
            # Restore buttons
            button_run.config(state=tk.NORMAL)
            button_cancel.config(state=tk.DISABLED)
            
            try: 
                word_app.Quit()
            except: 
                pass
            try: 
                excel_app.Quit()
            except: 
                pass

    atr_thread = threading.Thread(target=target, daemon=True)
    atr_thread.start()
            

# main window
root = tk.Tk()
root.title("Annual TIF Report Program")

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()


def center_window(win, width, height):
    """Resize + center the given Tk window."""
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = (sw - width) // 2
    y = (sh - height) // 2
    win.geometry(f"{width}x{height}+{x}+{y}")

# … in your code …
window_width  = 720
window_height = 580
center_window(root, window_width, window_height)

# window position/dimensions
# position_top = int(screen_height / 2 - 250)
# position_right = int(screen_width / 2 - 300)
# root.geometry(f"750x650+{position_right}+{position_top}")

# text box for the reporting year
text_frame = ttk.Frame(root)  # Use a frame to group widgets
text_frame.pack(pady=10)

label = ttk.Label(text_frame, text="Reporting Year:")  # Label next to entry
label.pack(side=tk.LEFT, padx=5)

date_entry = ttk.Entry(text_frame, width=10)
date_entry.pack(side=tk.LEFT)

label = tk.Label(text_frame, text="")
label.pack(side=tk.LEFT, padx=20)

label = ttk.Label(text_frame, text="Output Data Tables") 
label.pack(side=tk.LEFT, padx=5)

dt_ceckbox_var = tk.BooleanVar(value=False)
data_table_checkbox = ttk.Checkbutton(text_frame, variable=dt_ceckbox_var)
data_table_checkbox.pack(side=tk.LEFT)

label = tk.Label(text_frame, text="")
label.pack(side=tk.LEFT, padx=20)

label = ttk.Label(text_frame, text="PDF Merger")
label.pack(side=tk.LEFT, padx=5)

pdfm_checkbox_var = tk.BooleanVar(value=False)
pdfm_checkbox = ttk.Checkbutton(text_frame, variable=pdfm_checkbox_var)
pdfm_checkbox.pack(side=tk.LEFT)

# file frame
file_frame = tk.Frame(root)
file_frame.pack(pady=10)

file_button = ttk.Button(file_frame, text="Input File", command=open_input_file)
file_button.pack(side=tk.LEFT)
file_label = ttk.Label(file_frame, text="No file selected            ")
file_label.pack(side=tk.LEFT, padx=5)

template_button = ttk.Button(file_frame, text="Template File", command=open_template_file)
template_button.pack(side=tk.LEFT)
template_label = ttk.Label(file_frame, text="No file selected               ")
template_label.pack(side=tk.LEFT, padx=5)

att_frame = tk.Frame(root)
att_frame.pack(pady=10)

attB_button = ttk.Button(att_frame, text="Att B TIFCorp", command=open_attB_file)
attB_button.pack(side=tk.LEFT)
attB_label = ttk.Label(att_frame, text="No file selected  ")
attB_label.pack(side=tk.LEFT, padx=5)

attB2_button = ttk.Button(att_frame, text="Att B IJRL", command=open_attB2_file)
attB2_button.pack(side=tk.LEFT)
attB2_label = ttk.Label(att_frame, text="No file selected  ")
attB2_label.pack(side=tk.LEFT, padx=5)

attC_button = ttk.Button(att_frame, text="Att C TIFCorp", command=open_attC_file)
attC_button.pack(side=tk.LEFT)
attC_label = ttk.Label(att_frame, text="No file selected  ")
attC_label.pack(side=tk.LEFT, padx=5)

attC2_button = ttk.Button(att_frame, text="Att C IJRL", command=open_attC2_file)
attC2_button.pack(side=tk.LEFT)
attC2_label = ttk.Label(att_frame, text="No file selected  ")
attC2_label.pack(side=tk.LEFT, padx=5)


signed_frame = tk.Frame(root)
signed_frame.pack(pady=10)

bsigned_button = ttk.Button(signed_frame, text="Att B Signed", command=open_bsigned_file)
bsigned_button.pack(side=tk.LEFT)
bsigned_label = ttk.Label(signed_frame, text="No file selected            ")
bsigned_label.pack(side=tk.LEFT, padx=5)

csigned_button = ttk.Button(signed_frame, text="Att C Signed", command=open_csigned_file)
csigned_button.pack(side=tk.LEFT)
csigned_label = ttk.Label(signed_frame, text="No file selected               ")
csigned_label.pack(side=tk.LEFT, padx=5)

# sorting frame
sort_frame = tk.Frame(root)
sort_frame.pack(pady=10)

# Create a frame to hold the canvas and scrollbar
frame = tk.Frame(root)
frame.pack(fill=tk.NONE, expand=False)

# Create a canvas widget
canvas = tk.Canvas(frame)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Scrollbar and mouse scrolling
scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

canvas.configure(yscrollcommand=scrollbar.set)
scrollbar.configure(command=canvas.yview)

def on_mouse_wheel(event):
    canvas.yview_scroll(-1*(event.delta//120), "units")
canvas.bind_all("<MouseWheel>", on_mouse_wheel)

# Create a frame to hold the content inside the canvas
content_frame = tk.Frame(canvas)

tif_inst = TIF_List()

def create_bordered_frame(parent, color, row, col):
    frame = tk.Frame(parent, bd=1, relief="solid", bg=color)
    frame.grid(row=row, column=col, padx=0, pady=5, sticky="nsew")
    return frame

def shift_click(event, index):
    if event.state & 0x0001:  # Check if Shift key is held down (bitmask)
        if tif_inst.last_clicked is not None:
            start, end = sorted([tif_inst.last_clicked, index])
            # Set all checkboxes in the range to checked (True)
            for i in range(start, end):
                tif_inst.checkboxes[i].set(True)  # Ensure all are checked
    tif_inst.last_clicked = index

def update_grid():    
    for widget in content_frame.winfo_children():
        widget.destroy()
        
    # Create grid (4 columns for checkbox, num, name, group)
    for i in range(tif_inst.size()):
        # Create StringVar for dynamic updates
        num_var = tk.StringVar(value=f"{tif_inst.list[i].num}")
        name_var = tk.StringVar(value=f"{tif_inst.list[i].name}")
        group_var = tk.StringVar(value=f"{tif_inst.list[i].group}")

        # Create the frames for each column of data
        num_frame = tk.Frame(content_frame)
        num_frame.grid(row=i, column=1, sticky="nsew")

        name_frame = tk.Frame(content_frame)
        name_frame.grid(row=i, column=2, sticky="nsew")

        group_frame = tk.Frame(content_frame)
        group_frame.grid(row=i, column=3, sticky="nsew")

        # Create the checkboxes and bind the shift-click event
        checkbox = ttk.Checkbutton(content_frame, variable=tif_inst.checkboxes[i])
        checkbox.grid(row=i, column=0, padx=1, pady=3, sticky="nsew")
        checkbox.bind("<Button-1>", lambda e, index=i: shift_click(e, index))

        # Create the labels for num, name, group using StringVar
        tk.Label(num_frame, textvariable=num_var, font=("Arial", 12)).grid(row=i, column=1, sticky="nsew")
        tk.Label(name_frame, textvariable=name_var, font=("Arial", 12)).grid(row=i, column=2, sticky="nsew")
        tk.Label(group_frame, textvariable=group_var, font=("Arial", 12)).grid(row=i, column=3, sticky="nsew")
        

    # Configure the grid to expand properly
    content_frame.grid_rowconfigure(0, weight=1)
    for i in range(tif_inst.size()):
        content_frame.grid_rowconfigure(i+1, weight=1)  # Allow rows to expand
    content_frame.grid_columnconfigure(0, weight=1)  # Allow columns to expand
    content_frame.grid_columnconfigure(1, weight=1)
    content_frame.grid_columnconfigure(2, weight=1)
    content_frame.grid_columnconfigure(3, weight=1)

    # Add the content_frame to the canvas window
    canvas.create_window((0, 0), window=content_frame, anchor="nw")

    # Update the scrollable region
    content_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

# Sorting functions tied to buttons
def sort_by_num():
    tif_inst.sort_by_column(1)
    update_grid()

def sort_by_name():
    tif_inst.sort_by_column(2)
    update_grid()

def sort_by_group():
    tif_inst.sort_by_column(3)
    update_grid()
    
def select_all():
    for var in tif_inst.checkboxes:
        var.set(True)
    update_grid()
    
def deselect_all():
    for var in tif_inst.checkboxes:
        var.set(False)
    update_grid()
    

# Create sorting buttons
button_num = ttk.Button(sort_frame, text="Sort by Num", command=sort_by_num)
button_num.pack(side="left", padx=5)

button_name = ttk.Button(sort_frame, text="Sort by Name", command=sort_by_name)
button_name.pack(side="left", padx=5)

button_group = ttk.Button(sort_frame, text="Sort by Group", command=sort_by_group)
button_group.pack(side="left", padx=5)

label_spacer = ttk.Label(sort_frame, text="                           ")
label_spacer.pack(side=tk.LEFT)

button_num = ttk.Button(sort_frame, text="Select All", command=select_all)
button_num.pack(side="right", padx=5)

button_num = ttk.Button(sort_frame, text="Deselect All", command=deselect_all)
button_num.pack(side="right", padx=5)

# run frame
run_frame = tk.Frame(root)
run_frame.pack(pady=10)

button_run = ttk.Button(run_frame, text="Run", command=run_code)
button_run.pack(side=tk.LEFT)

run_buf = ttk.Label(run_frame, text="                  ")
run_buf.pack(side=tk.LEFT)

button_cancel = ttk.Button(run_frame, text="Cancel", command=cancel_run, state=tk.DISABLED)
button_cancel.pack(side=tk.LEFT)

run_status_frame = tk.Frame(root)
run_status_frame.pack(pady=10)

run_status = ttk.Label(run_status_frame, text="")
run_status.pack(side=tk.LEFT)

# Initial grid display
update_grid()

load_state()

# cProfile.run('update_grid()')

root.protocol("WM_DELETE_WINDOW", lambda: (save_state(), root.destroy()))

# Run the GUI
root.mainloop()

    

