from openpyxl import load_workbook, Workbook
import shutil
from copy import copy
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime
from datetime import date
import os
import sys
import win32com.client
import re
from pathlib import Path
from PDF_Merger import merge_by_tif_number, bc_docs
from pypdf import PdfReader, PdfWriter
import tkinter
from tkinter import messagebox


def column_match(label_row: int, data, column_labels: map):
    """Matches labels to their columns in the excel file

    Args:
        label_row (int): row that holds the labels of the columns (in data)
        data: sheet that contains the input data
        column_labels (map): mapped list that contains the lowercase names of important columns

    Returns:
        column_labels: updated map of column lables
    """
    for cell in data[label_row]:
        if cell.value:
            label = str(cell.value).strip().lower()
            if label in column_labels:
                column_labels[label] = cell.column
    return column_labels

def get_tif_rows(row, num_col, data, current_num):
    """Finds the rows in an excel sheet that hold the data important to the current TIF

    Args:
        row (int): starting row for the input data
        num_col (int): the column that holds the TIF numbers
        data (_type_): data sheet input
        current_num (int): current TIF number to look for

    Returns:
        list: rows that contains data for the specified TIF
    """
    tif_row = []
    while True:
        num = normalized_value(data.cell(row=row, column=num_col).value)
        if num is None:
            break
        if num == current_num:
            tif_row.append(row)
        row += 1
    return tif_row

# variation of previous function for only a single row
def get_tif_row_single(row, num_col, data, current_num):
    """Finds the row in an excel sheet that holds the data important to the current TIF

    Args:
        row (int): starting row for the input data
        num_col (int): the column that holds the TIF numbers
        data (_type_): data sheet input
        current_num (int): current TIF number to look for

    Returns:
        int: row containing data for the specified TIF (-1 if not found)
    """
    while True:
        num = normalized_value(data.cell(row=row, column=num_col).value)
        if num is None:
            break
        if num == current_num:
            return row
        row += 1
    return -1
    
    
def sort_single(data, destination, row: int, label_row: int, column_labels: map, column_map: map, row_map: map, current_num: int):
    """
    Args:
        data: sheet that contains the input data
        destination: sheet that contains the output data
        row: starting row for the input data
        label_row: row that holds the labels of the columns (in data)
        column_labels: mapped list that contains the lowercase names of important columns
        column_map: mapped list that contains the location of the columns in the output
        row_map (list): mapped list that contains the location of the rows in the output
        current_num: current tif number to look for
        
    """
    # get input column numbers
    column_labels = column_match(label_row, data, column_labels)
    
    # finding TIF row
    tif_row = -1
    while True:
        num = normalized_value(data.cell(row=row, column=column_labels['tifnum']).value)
        if num is None:
            break
        if num == current_num:
            tif_row = row
        row += 1
        
    # make sure data for the current TIF exists
    # if tif_row == -1:
        
    #     print(f"TIF {current_num} was not found in section")
    #     return
    
    # sorting 
    for label, data_col in column_labels.items():
        if label == 'tifnum':
            continue
        
        col = column_map[label]
        row = row_map[label]
        
        value = normalized_value(data.cell(row=tif_row, column=data_col).value) if data_col >= 1 else None
        if value not in (0, None):
            destination.cell(row=row, column=col).value = value
    return

# data          = sheet that contains the input data
# destination   = sheet that contains the output data
# row           = starting row for the input data
# label_row     = row that holds the labels of the columns (in data)
# column_labels = mapped list that contains the lowercase names of important columns
# column_map    = mapped list that contains the location of the columns in the output
# row_map       = mapped list that contains the location of the rows in the output
# current_num   = current tif number to look for
def sort_single_prev(data, prev_data, destination, row, label_row, prev_label_row, column_labels, prev_labels, column_map, row_map, current_num):
    # get input column numbers
    column_labels = column_match(label_row, data, column_labels)
    
    # finding TIF row
    tif_row = -1
    while True:
        num = normalized_value(data.cell(row=row, column=column_labels['tifnum']).value)
        if num is None:
            break
        if num == current_num:
            tif_row = row
        row += 1
    
    prev_tif_row = -1
    while True:
        num = normalized_value(data.cell(row=row, column=column_labels['tifnum']).value)
        if num is None:
            break
        if num == current_num:
            prev_tif_row = row
        row += 1
        
    # make sure data for the current TIF exists
    # if tif_row == -1:
        
    #     print(f"TIF {current_num} was not found in ")
    #     return
    
    # sorting 
    for label, data_col in column_labels.items():
        if label == 'tifnum':
            continue
        
        col = column_map[label]
        row = row_map[label]
        
        value = normalized_value(data.cell(row=tif_row, column=data_col).value) if data_col >= 1 else None
        if value not in (0, None):
            destination.cell(row=row, column=col).value = value
    return


def sort_multiple(data, destination, row: int, label_row: int, column_labels: map, column_map: map, row_map: map, current_num: int):
    """Populates a sheet according to the mapping when a single TIF can have mutiple instances of data\n\n
    _WARNING:_ Don't use this function if data of subsequent tif instances don't go directly below previous data

    Args:
        data (_type_): sheet that contains the input data
        destination (_type_): sheet that contains the output data
        row (int): starting row for the input data
        label_row (map): row that holds the labels of the columns (in data)
        column_labels (map): mapped list that contains the lowercase names of important columns
        column_map (map): mapped list that contains the location of the columns in the output
        row_map (map): mapped list that contains the location of the rows in the output
        current_num (int): current tif number to look for
    """
    
    # get input column numbers
    column_labels = column_match(label_row, data, column_labels)
    
    # finding TIF rows
    tif_rows = []
    while True:
        num = normalized_value(data.cell(row=row, column=column_labels['tifnum']).value)
        if num is None:
            break
        if num == current_num:
            tif_rows.append(row)
        row += 1
    
    if len(tif_rows) == 0:
        return
    
    # sorting
    for i in range(len(tif_rows)):
        for label, data_col in column_labels.items():
            if label == 'tifnum':
                continue
            
            col = column_map[label]
            row = row_map[label]
            
            value = normalized_value(data.cell(row=tif_rows[i], column=data_col).value) if data_col >= 1 else None
            if value != None:
                destination.cell(row=row+i, column=col).value = value
    return

SECTION_RE = re.compile(r"^\s*(section|attachment)\s+(.+)$", re.I)

def sheet_suffix(sheet_name: str) -> str:
    """
    Convert worksheet titles like
        'Section 1'        -> '1'
        'Section 3.2 A'    -> '3.2A'
        'ATTACHMENT E'     -> 'E'
    Anything that doesn’t start with Section/Attachment is returned unchanged.
    """
    m = SECTION_RE.match(sheet_name)
    suffix = m.group(2) if m else sheet_name          # part after the keyword
    return suffix.replace(" ", "")                    # kill internal spaces

def normalized_value(v):
    """
    Read an openpyxl cell and return None if
    its .value is None or if it’s a str of only whitespace.
    Otherwise return the original .value (with stripped whitespace for strings).
    """
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v



def ATR(tif_list, section1_list, reporting_year, input_file, template_file, attB_tifcorp_file, attB_ijrl_file, 
        attC_tifcorp_file, attC_ijrl_file, bsigned_file, csigned_file, merge):
    
    def section_1(sec_1_list):
        destination = current_ATR['Section 1']
        data = Data_Tables['Section 1']
        
        footnote_idx = 44
        # Read the existing footnote cells & row height
        footnote_cells = list(destination[footnote_idx])
        footnote_height = destination.row_dimensions[footnote_idx].height
        # Remove that row so everything shifts up
        destination.delete_rows(footnote_idx, 1)
        
        
        # Part 1: extending rows to reach the desired length
        row_idx = 43
        next_row_idx = row_idx + 1
        
        base_cells = list(destination[row_idx])
        base_height = destination.row_dimensions[row_idx].height

        for _ in range(len(sec_1_list) - 15):
            destination.insert_rows(next_row_idx)  # Insert a new row before the next row
            destination.row_dimensions[next_row_idx].height = base_height
            
            destination.merge_cells(start_row=next_row_idx, start_column=2, end_row=next_row_idx, end_column=6)
            destination.merge_cells(start_row=next_row_idx, start_column=7, end_row=next_row_idx, end_column=8)
            destination.merge_cells(start_row=next_row_idx, start_column=9, end_row=next_row_idx, end_column=10)
            
            for col_idx, cell in enumerate(base_cells, start=1):
                new_cell = destination.cell(row=next_row_idx, column=col_idx, value=cell.value)

                # Copy formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.protection = copy(cell.protection)
            next_row_idx += 1
        
        destination.insert_rows(next_row_idx)
        destination.row_dimensions[next_row_idx].height = footnote_height
        destination.merge_cells(start_row=next_row_idx, start_column=2, end_row=next_row_idx, end_column=10)

        # Restore footnote cell values & formatting
        for col_idx, old_cell in enumerate(footnote_cells, start=1):
            new_cell = destination.cell(row=next_row_idx, column=col_idx, value=old_cell.value)
            if old_cell.has_style:
                new_cell.font = copy(old_cell.font)
                new_cell.border = copy(old_cell.border)
                new_cell.fill = copy(old_cell.fill)
                new_cell.number_format = copy(old_cell.number_format)
                new_cell.alignment = copy(old_cell.alignment)
                new_cell.protection = copy(old_cell.protection)
        
        destination.unmerge_cells(start_row=row_idx + 1, start_column=2, end_row=row_idx + 1, end_column=10)
        destination.merge_cells(start_row=row_idx + 1, start_column=2, end_row=row_idx + 1, end_column=6)
        destination.merge_cells(start_row=row_idx + 1, start_column=7, end_row=row_idx + 1, end_column=8)
        destination.merge_cells(start_row=row_idx + 1, start_column=9, end_row=row_idx + 1, end_column=10)
        
        for col_idx, cell in enumerate(base_cells, start=1):
                new_cell = destination.cell(row=row_idx + 1, column=col_idx, value=cell.value)

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.protection = copy(cell.protection)
        
        # Part 2: Filling data
        column_labels = {
            'tifnum': None, 'tifname': None, 'approvedate': None, 'expiredate': None,
        }
        column_map = {
            'tifname': 2, 'approvedate': 7, 'expiredate': 9
        }

        # Locate column indices
        column_labels = column_match(label_row=1, data=data, column_labels=column_labels)

        # Choose the column to sort by – for example, sorting by 'tifname'
        sort_label = 'tifname'
        sort_index = column_labels[sort_label] - 1  # Convert to 0-based index

        # Read all data rows into a list and sort them by the chosen column
        all_rows = list(data.iter_rows(min_row=2, values_only=False))
        sorted_rows = sorted(
            all_rows,
            key=lambda row: (row[sort_index].value or "").lower()  # lower() ensures case-insensitive sorting
        )

        row_check = 29  # Change if start of data position changes
        d_row = row_check
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # highlighter color

        # Iterate over the sorted rows instead of the unsorted iterator
        for row_cells in sorted_rows:
            row_values = [cell.value for cell in row_cells]
            
            for label, dest_col  in column_map.items():
                col_index = column_labels[label] - 1
                value = row_values[col_index] if col_index >= 0 else None
                if value not in (0, None):
                    if label == 'expiredate':

                        
                        dest_cell = destination.cell(row=d_row, column=dest_col, value=value)
                        dest_cell.number_format = 'MM/DD/YYYY'

                        if isinstance(value, str):
                            try:
                                from datetime import datetime
                                value = datetime.strptime(value, "%m/%d/%Y")
                                # Optionally update the cell with the converted value:
                                dest_cell.value = value
                            except ValueError:
                                continue  # Skip this cell if conversion fails

                        # Check the year AFTER conversion
                        if value.year == reporting_year_int:
                            dest_cell.fill = yellow_fill  # Highlight cell
                    else:
                        dest_cell = destination.cell(row=d_row, column=dest_col, value=value)
                    if label == 'tifname' and value == current_name:
                        destination.cell(row=d_row, column=1, value="X")
            d_row += 1
        
        nonlocal footnote_1, footnote_3, footnote_6, footnote_7, fn6_text, fn7_text, asterisk_check
        for i in range(len(sec_1_list)):
            tif_name_cell = destination.cell(row=row_check + i, column=column_map['tifname'])
            tif_name = tif_name_cell.value
            if tif_name == current_name:
                tif_row = tif_name_cell.row
                tif_app_date = destination.cell(row=tif_row, column=column_map['approvedate']).value
                tif_exp_date = destination.cell(row=tif_row, column=column_map['expiredate']).value
                
                if tif_app_date.year < 2003:
                    footnote_1 = True
                    
                date_string = "11/1/1999"
                from datetime import datetime
                date_obj = datetime.strptime(date_string, "%m/%d/%Y")
                if tif_app_date < date_obj:
                    print("tif is older than 1999")
                    footnote_3 = True
                    
                if tif_exp_date.year == reporting_year_int:
                    print("year is terminating")
                    asterisk_check = True
                    footnote_6 = True
                    fn6_text = f"** The Project terminated on {tif_exp_date.month}/{tif_exp_date.day}/{tif_exp_date.year}.  This line equals the net {reporting_year} surplus, which has been settled."
                    footnote_7 = True
                    fn7_text = f"* The Project terminated on {tif_exp_date.month}/{tif_exp_date.day}/{tif_exp_date.year}.  Therefore, this line equals the Fund Balance by Source as of the termination date of the Project."
            
        
        # nonlocal 
        # tif_name = destination.cell(row=offset_final, column=column_map['tifname']).value
        # print(f"{tif_name}, {current_name}")
        # if tif_name == current_name:
        #     foot_value = destination.cell(row=offset_final, column=column_map['expiredate']).value
        #     print(f"foot_value: {foot_value}")
        #     if isinstance(foot_value, date):
        #         print("foot is a date")
        #         if foot_value.year == reporting_year_int:
        #             print("year is terminating")
        #             footnote_6 = True
        #             fn6_text = f"** The Project terminated on {foot_value}.  This line equals the net {reporting_year} surplus, which has been settled."
        #             footnote_7 = True
        return
    
    def section_2():
        destination = current_ATR['Section 2']
        data = Data_Tables['Section 2']
        
        destination.cell(row=4, column=1, value=current_name)
        column_labels = {
            'tifnum': None, 'primaryuse': None, 'combomix': None, 'ijrl': None
        }
        
        column_labels = column_match(label_row=1, data=data, column_labels=column_labels)
        tif_row = get_tif_row_single(row=2, num_col=column_labels['tifnum'], data=data, current_num=current_num)
        
        destination.cell(row=6, column=2).value = normalized_value(data.cell(row=tif_row, column=column_labels['primaryuse']).value)
        
        value = destination.cell(row=8, column=1).value
        new_value = normalized_value(data.cell(row=tif_row, column=column_labels['combomix']).value)
        if new_value not in (0, None):
            value = value + f"{new_value}"
            destination.cell(row=8, column=1, value=value)
        
        nonlocal ijrl
        value = normalized_value(data.cell(row=tif_row, column=column_labels['ijrl']).value)
        if value:
            print("entered ijrl")
            ijrl = True
            destination.cell(row=11, column=2, value="__X___")
        else:
            destination.cell(row=10, column=2, value="__X___")
            
        # Attachment A
        data = Data_Tables['Attachment A']
        a_labels = {
            'tifnum': None
        }
        a_labels = column_match(label_row=1, data=data, column_labels=a_labels)
        tif_row = get_tif_row_single(row=2, num_col=a_labels['tifnum'], data=data, current_num=current_num)
        if tif_row > 0:
            cell = destination.cell(row=15, column=3, value='X')
            cell.font = Font(bold=True)
            # attach files
        else:
            cell = destination.cell(row=15, column=2, value='X')
            cell.font = Font(bold=True)
        
        # Attachment D
        data = Data_Tables['Attachment D']
        destination = current_ATR['ATTACHMENT D']
        d_labels = {
            'tifnum': None, 'projectname': None
        }
        d_labels = column_match(label_row=1, data=data, column_labels=d_labels)
        tif_rows = get_tif_rows(row=2, num_col=d_labels['tifnum'], data=data, current_num=current_num)
        
        if len(tif_rows) == 0:
            del current_ATR['ATTACHMENT D']
            destination = current_ATR['Section 2']
            destination.cell(row=18, column=2, value='X')
        else:
            if len(tif_rows) > 1:
                row_idx = 10
                next_row_idx = row_idx + 1
                row_cells = [destination[row_idx][col] for col in range(len(destination[row_idx]))]  # Get original row cells
                row_height = destination.row_dimensions[row_idx].height  # Save row height
                
                for _ in range(len(tif_rows) - 1):
                    destination.insert_rows(next_row_idx)  # Insert a new row before the next row
                    destination.merge_cells(start_row=next_row_idx, start_column=1, end_row=next_row_idx, end_column=5)
                    destination.row_dimensions[next_row_idx].height = row_height
                    
                    for col_idx, cell in enumerate(row_cells, start=1):
                        new_cell = destination.cell(row=next_row_idx, column=col_idx, value=cell.value)

                        # Copy formatting
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.protection = copy(cell.protection)
                    next_row_idx += 1
            starting_row = 10
            for i in range(len(tif_rows)):
                destination.cell(row=starting_row+i, column=1).value = normalized_value(data.cell(row=tif_rows[i], column=d_labels['projectname']).value)
            
            destination = current_ATR['Section 2']
            cell = destination.cell(row=18, column=3, value='X') 
            cell.font = Font(bold=True)               
            
        
        # Attachment E
        data = Data_Tables['Attachment E']
        destination = current_ATR['ATTACHMENT E']
        e_labels = {
            'tifnum': None, 'address': None, 'project description': None
        }
        e_labels = column_match(label_row=1, data=data, column_labels=e_labels)
        tif_rows = get_tif_rows(row=2, num_col=e_labels['tifnum'], data=data, current_num=current_num)
        
        if len(tif_rows) == 0:
            del current_ATR['ATTACHMENT E']
            destination = current_ATR['Section 2']
            destination.cell(row=19, column=2, value='X')
        else:
            if len(tif_rows) > 1:
                row_idx = 12
                next_row_idx = row_idx + 1
                row_cells = [destination[row_idx][col] for col in range(len(destination[row_idx]))]  # Get original row cells
                row_height = destination.row_dimensions[row_idx].height  # Save row height
                
                for _ in range(len(tif_rows) - 1):
                    destination.insert_rows(next_row_idx)  # Insert a new row before the next row
                    destination.merge_cells(start_row=next_row_idx, start_column=1, end_row=next_row_idx, end_column=3)
                    destination.merge_cells(start_row=next_row_idx, start_column=4, end_row=next_row_idx, end_column=5)
                    destination.merge_cells(start_row=next_row_idx, start_column=6, end_row=next_row_idx, end_column=7)
                    destination.row_dimensions[next_row_idx].height = row_height
                    
                    for col_idx, cell in enumerate(row_cells, start=1):
                        new_cell = destination.cell(row=next_row_idx, column=col_idx, value=cell.value)

                        # Copy formatting
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.protection = copy(cell.protection)
                    next_row_idx += 1
            starting_row = 12
            for i in range(len(tif_rows)):
                destination.cell(row=starting_row+i, column=1, value="N/A")
                destination.cell(row=starting_row+i, column=4).value = normalized_value(data.cell(row=tif_rows[i], column=e_labels['project description']).value)
                destination.cell(row=starting_row+i, column=6).value = normalized_value(data.cell(row=tif_rows[i], column=e_labels['address']).value)
            
            destination = current_ATR['Section 2']
            cell = destination.cell(row=19, column=3, value='X') 
            cell.font = Font(bold=True)
                        
                        
        # Attachment F
        cell = destination.cell(row=20, column=2, value='X')
        cell.font = Font(bold=True)
        
        
        # Attachment G
        data = Data_Tables['Attachment G']
        g_labels = {
            'tifnum': None
        }
        g_labels = column_match(label_row=1, data=data, column_labels=g_labels)
        tif_rows = get_tif_rows(row=2, num_col=g_labels['tifnum'], data=data, current_num=current_num)
        
        if len(tif_rows) > 0:
            cell = destination.cell(row=21, column=3, value='X')
            cell.font = Font(bold=True)
            # attach files
        else:
            cell = destination.cell(row=21, column=2, value='X')
            cell.font = Font(bold=True)
        
        # Attachment H
        data = Data_Tables['Attachment H']        
        h_labels = {
            'tifnum': None
        }
        h_labels = column_match(label_row=1, data=data, column_labels=h_labels)
        tif_rows = get_tif_rows(row=2, num_col=h_labels['tifnum'], data=data, current_num=current_num)
        
        if len(tif_rows) > 0:
            cell = destination.cell(row=22, column=3, value='X')
            cell.font = Font(bold=True)
            # attach files
        else:
            cell = destination.cell(row=22, column=2, value='X')
            cell.font = Font(bold=True)
        
        # Attachment I
        cell = destination.cell(row=23, column=2, value='X')
        cell.font = Font(bold=True)
        
        # Attachment J
        cell = destination.cell(row=24, column=2, value='X')
        cell.font = Font(bold=True)
        
        # Attachment K
        # if attach_K == True:
        #     cell = destination.cell(row=25, column=3, value='X')
        #     cell.font = Font(bold=True)
        # else:
        #     cell = destination.cell(row=25, column=2, value='X')
        #     cell.font = Font(bold=True)
            
        # # Attachment L
        # if attach_L == True:
        #     cell = destination.cell(row=26, column=3, value='X')
        #     cell.font = Font(bold=True)
        # else:
        #     cell = destination.cell(row=26, column=2, value='X')
        #     cell.font = Font(bold=True)
            
        # Attachment M covered in section 5
        
        # Attachment N
        cell = destination.cell(row=28, column=2, value='X')
        cell.font = Font(bold=True)
        return
    
    def section_3_1():
        def section_3_1_Other():
            destination = current_ATR['Section 3.1 Other']
            data = Data_Tables['Section 3.1 Other']
            prev_data = Data_Tables['Section 3.1 Other Previous']
            
            column_labels = {
                'tifnum': None, 'noteproceedscurrentyear': None, 'noncompliancepayment': None, 
                'excessreserverequirement': None, 'babrebate': None, 'collectionreturns': None, 
                'creditsexpenditures': None, 'prioryearscumulative': None # added or else excluded
            }
            mapping = {
                'noteproceedscurrentyear': 'noteproceedscumulative', 'noncompliancepayment': 'noncompliancecum', 
                'excessreserverequirement': 'excessreservecum', 'babrebate': 'babrebatecum', 'collectionreturns': 
                'collectionreturnscum', 'creditsexpenditures': 'creditsexpenditurescum'
            }
            prev_labels = {
                'tifnum': None, 'prioryearscumulative': None, 'noteproceedscumulative': None, 
                'noncompliancecum': None, 'excessreservecum': None, 'babrebatecum': None, 
                'collectionreturnscum': None, 'creditsexpenditurescum': None
            }
            column_map = {
                'prioryearscumulative': 7, 'noteproceedscurrentyear': 6, 'noteproceedscumulative': 7, 
                'noncompliancepayment': 6, 'noncompliancecum': 7, 'excessreserverequirement': 6, 
                'excessreservecum': 7, 'babrebate': 6, 'babrebatecum': 7, 'collectionreturns': 6, 
                'collectionreturnscum': 7, 'creditsexpenditures': 6, 'creditsexpenditurescum': 7
            }
            row_map = {
                'prioryearscumulative': 10, 'noteproceedscurrentyear': 11, 'noteproceedscumulative': 11, 
                'noncompliancepayment': 12, 'noncompliancecum': 12, 
                'excessreserverequirement': 13, 'excessreservecum': 13, 'babrebate': 14, 'babrebatecum': 14, 
                'collectionreturns': 15, 'collectionreturnscum': 15, 
                'creditsexpenditures': 16, 'creditsexpenditurescum': 16
            }
            
            label_row = 1
            column_labels = column_match(label_row, data, column_labels)

            label_row = 1
            prev_labels = column_match(label_row, prev_data, prev_labels)
            
            # finding TIF row
            row = 4
            tif_row = get_tif_row_single(row=row, num_col=column_labels['tifnum'], data=data, current_num=current_num)
            
            row = 2
            prev_tif_row = get_tif_row_single(row=row, num_col=prev_labels['tifnum'], data=prev_data, current_num=current_num)
                
            # make sure data for the current TIF exists
            # if tif_row == -1 or prev_tif_row == -1:
            #     if tif_row == -1:
            #         print(f"TIF {current_num} was not found in Section 3.1 regular")
            #     if prev_tif_row == -1:
            #         print(f"TIF {current_num} was not found in Section 3.1 Previous")
            #     return

            # sorting 
            for label, data_col in column_labels.items():
                if label == 'tifnum':
                    continue
                    
                if label == 'prioryearscumulative':
                    row = row_map['prioryearscumulative']
                    col = column_map['prioryearscumulative']
                    value = normalized_value(prev_data.cell(row=prev_tif_row, column=prev_labels['prioryearscumulative']).value)
                    if value not in (0, None):
                        destination.cell(row=row, column=col).value = value
                        sheets_to_skip['section 3.1 other'] = False
                    continue
                
                col = column_map[label]
                row = row_map[label]
                
                value = normalized_value(data.cell(row=tif_row, column=data_col).value) if data_col >= 1 else None
                value = int(value) if isinstance(value, (int, float)) else 0
                if value not in (0, None):
                    destination.cell(row=row, column=col).value = value
                    print(f"{label}, {row}, {col}")
                    sheets_to_skip['section 3.1 other'] = False
                    
                if label in mapping:
                    prev_label = mapping[label]
                    col = column_map[prev_label]
                    row = row_map[prev_label]
                    d_col = prev_labels[prev_label]
                    
                    prev_value = normalized_value(prev_data.cell(row=prev_tif_row, column=d_col).value) if data_col >= 1 else None
                    prev_value = int(prev_value) if isinstance(prev_value, (int, float)) else 0
                    prev_value += value
                    if prev_value not in (0, None):
                        print(prev_value)
                        destination.cell(row=row, column=col).value = prev_value
                        sheets_to_skip['section 3.1 other'] = False
        

        destination = current_ATR['Section 3.1']
        data = Data_Tables['Section 3.1']
        prev_data = Data_Tables['Section 3.1 Previous']
        
        column_labels = {
            'tifnum': None, 'taxallocationfundbalance': None, 'proptaxincr-current': None, 
            'interest-current': None, 'land/bldg-current': None, 'bond-current': None, 
            'municipal-current': None, 'private-current': None, 'totalexp/cash': None, 
            'distributionofsurplus': None, 'transfers--municipal': None
        }
        mapping = {
            'proptaxincr-current': 'proptaxincr-previous',
            'interest-current': 'interest previous',
            'land/bldg-current': 'land building sale previous',
            'bond-current': 'bond proceeds previous',
            'municipal-current': 'transfers to municipal sources previous',
            'private-current': 'private sources previous'
        }
        prev_labels = {
            'tifnum': None, 'proptaxincr-previous': None, 'interest previous': None, 
            'land building sale previous': None, 'bond proceeds previous': None,
            'note previous': None, 'transfers to municipal sources previous': None, 
            'private sources previous': None
        }
        column_map = {
            'taxallocationfundbalance': 2, 'proptaxincr-current': 2, 'proptaxincr-previous': 3,
            'interest-current': 2, 'interest previous': 3, 'land/bldg-current': 2,
            'land building sale previous': 3, 'bond-current': 2, 'bond proceeds previous': 3,
            'municipal-current': 2, 'transfers to municipal sources previous': 3, 'private-current': 2, 
            'private sources previous': 3, 'totalexp/cash': 2, 'distributionofsurplus': 2,
            'transfers--municipal': 2
        }
        row_map = {
            'taxallocationfundbalance': 8, 'proptaxincr-current': 11, 'proptaxincr-previous': 11,
            'interest-current': 16, 'interest previous': 16, 'land/bldg-current': 17,
            'land building sale previous': 17, 'bond-current': 18, 'bond proceeds previous': 18,
            'municipal-current': 19, 'transfers to municipal sources previous': 19, 'private-current': 20,
            'private sources previous': 20, 'totalexp/cash': 27, 'distributionofsurplus': 29,
            'transfers--municipal': 28
        }
        
        label_row = 1
        column_labels = column_match(label_row, data, column_labels)

        label_row = 1
        prev_labels = column_match(label_row, prev_data, prev_labels)
        
        # finding TIF row
        row = 4
        tif_row = -1
        while True:
            num = normalized_value(data.cell(row=row, column=column_labels['tifnum']).value)
            if num is None:
                break
            if num == current_num:
                tif_row = row
            row += 1
        
        row = 3
        prev_tif_row = -1
        while True:
            num = normalized_value(prev_data.cell(row=row, column=prev_labels['tifnum']).value)
            if num is None:
                break
            if num == current_num:
                prev_tif_row = row
            row += 1
            
        # make sure data for the current TIF exists
        # if tif_row == -1 or prev_tif_row == -1:
        #     if tif_row == -1:
        #         print(f"TIF {current_num} was not found in Section 3.1 regular")
        #     if prev_tif_row == -1:
        #         print(f"TIF {current_num} was not found in Section 3.1 Previous")
        #     return
        
        # sorting 
        for label, data_col in column_labels.items():
            if label == 'tifnum':
                print("TIFnum skipped")
                continue
                
            col = column_map[label]
            row = row_map[label]
            
            value = normalized_value(data.cell(row=tif_row, column=data_col).value) if data_col >= 1 else None
            value = int(value) if isinstance(value, (int, float)) else 0
            if value not in (0, None):
                destination.cell(row=row, column=col).value = value
                
            if label in mapping:
                prev_label = mapping[label]
                col = column_map[prev_label]
                row = row_map[prev_label]
                d_col = prev_labels[prev_label]
                
                prev_value = normalized_value(prev_data.cell(row=prev_tif_row, column=d_col).value) if data_col >= 1 else None
                prev_value = int(prev_value) if isinstance(prev_value, (int, float)) else 0
                prev_value += value
                if prev_value not in (0, None):
                    destination.cell(row=row, column=col).value = prev_value
        
        nonlocal footnote_1, footnote_6, fn6_text, asterisk_check #attach_K, attach_L
        if footnote_1:
            footnote = "(a) Cumulative figures for the categories of 'Interest,' 'Land/Building Sale Proceeds' and 'Other' may not be fully available for this report due to either: (i) the disposal of certain older records pursuant to the City's records retention policy, or (ii) the extraordinary administrative burden of developing cumulative City records prior to the City's conversion to its current accounting system in 2003."
            destination.cell(row=49, column=1, value=footnote)
        if footnote_6:
            footnote = fn6_text
            destination.cell(row=51, column=1, value=footnote)
        if asterisk_check:
            destination.cell(row=29, column=3, value="**")
            
        # if destination.cell(row=11, column=3).value > 100000:
        #     attach_K = True
        # if destination.cell(row=11, column=3).value > 100000:
        #     attach_L = True
            
        section_3_1_Other()

        
                
    def section_3_2_A():    
        destination = current_ATR['Section 3.2 A']
        data = Data_Tables['Section 3.2a']
        
        column_labels = {
            'tifnum': None, 'costofstudies': None, 'administrative cost': None, 
            'marketing sites': None, 'site preparation costs': None, 'renovation rehab, etc': None, 
            'public works': None, 'removing contaminants': None, 'jobtraining': None, 
            'financing costs': None, 'capital costs': None, 'schooldistricts': None, 
            'librarydistricts': None, 'relocation costs': None, 'inlieu of taxes': None, 
            'jobtraining-retraining2': None, 'interest cost': None, 'newhousing': None, 
            'daycare services': None, 'other': None
        }
        column_map = {
            'costofstudies': 3, 'administrative cost': 3, 'marketing sites': 3, 
            'site preparation costs': 3, 'renovation rehab, etc': 3, 'public works': 3, 
            'removing contaminants': 3, 'jobtraining': 3, 'financing costs': 3, 
            'capital costs': 3, 'schooldistricts': 3, 'librarydistricts': 3, 
            'relocation costs': 3, 'inlieu of taxes': 3, 'jobtraining-retraining2': 3, 
            'interest cost': 3, 'newhousing': 3, 'daycare services': 3, 'other': 3
        }
        row_map = {
            'costofstudies': 10, 'administrative cost': 18, 'marketing sites': 26, 
            'site preparation costs': 34, 'renovation rehab, etc': 42, 'public works': 50, 
            'removing contaminants': 60, 'jobtraining': 68, 'financing costs': 76, 
            'capital costs': 84, 'schooldistricts': 92, 'librarydistricts': 100, 
            'relocation costs': 110, 'inlieu of taxes': 118, 'jobtraining-retraining2': 126, 
            'interest cost': 134, 'newhousing': 135, 'daycare services': 142, 'other': 150
        }
        
        label_row = 1
        starting_row = 5
        
        column_labels = column_match(label_row, data, column_labels)
        tif_row = get_tif_row_single(starting_row, column_labels['tifnum'], data, current_num)
        sort_single(data, destination, starting_row, tif_row, column_labels, column_map, row_map, current_num)
        
        destination.cell(row=row_map['interest cost'], column=1).value = 'Costs of interest incurred by a Developer related to the construction, renovation or rehabilitation'
        destination.cell(row=row_map['newhousing'], column=1).value = 'Costs of construction of new housing units for low income or very low income households.'
        return
        
    def section_3_2_B():    
        destination = current_ATR['Section 3.2 B']
        data = Data_Tables['Section 3.2b']
        
        class vendor:
            def __init__(self, data1=None, data2=None, data3=None):
                self.name = data1
                self.service = data2
                self.amount = data3
        
        row = 2
        match = 0
        vendors = []
        while True:
            num = normalized_value(data.cell(row=row, column=1).value)
            if num is None:
                break
            if num == current_num:
                name = normalized_value(data.cell(row=row, column=3).value)
                service = normalized_value(data.cell(row=row, column=4).value)
                amount = normalized_value(data.cell(row=row, column=5).value)
                
                vendor_inst = vendor(name, service, amount)
                vendors.append(vendor_inst)
                match = 1
            row += 1
        
        nonlocal footnote_2    
        if match == 1:
            vendors.sort(key=lambda x: (x.name.strip().lower() != "city staff costs", x.service.lower(), x.name.lower()))
            for vendor_inst in vendors:
                if vendor_inst.name.strip().lower() == "city staff costs":
                    vendor_inst.name = vendor_inst.name.strip() + " (1)"
                    footnote_2 = True
                    break
            row = 9
            for i in range(len(vendors)):
                destination.cell(row=row+i, column=1).value = vendors[i].name
                destination.cell(row=row+i, column=2).value = vendors[i].service
                destination.cell(row=row+i, column=3).value = vendors[i].amount
        
        nonlocal footnote_3
        if footnote_2:
            footnote = "(1) Costs relate directly to the salaries and fringe benefits of employees working solely on tax increment financing districts."
            destination.cell(row=42, column=1, value=footnote)
        if footnote_3:
            footnote = "* This table may include payments for Projects that were undertaken prior to 11/1/1999."
            destination.cell(row=43, column=1, value=footnote)
        return
       
    def section_3_3():    
        destination = current_ATR['Section 3.3']
        data = Data_Tables['Section 3.3']
        
        
        column_labels = {
            'tifnum': None, 'descriptions of project costs to be paid': None, 
            'amount designated _(project costs)': None
        }
        column_map = {
            'descriptions of project costs to be paid': 1, 'amount designated _(project costs)': 3,
        }
        row_map = {
            'descriptions of project costs to be paid': 24, 'amount designated _(project costs)': 24,
        }
        
        label_row = 1
        starting_row = 4
        
        sort_multiple(data, destination, starting_row, label_row, column_labels, column_map, row_map, current_num)
        
        column_labels = {
            'tifnum': None, 'description of debt obligations': None, 'amount of original issuance': None,
            'amount designated (debt obligations)': None
        }
        column_map = {
            'description of debt obligations': 1, 'amount designated (debt obligations)': 3,
            'amount of original issuance': 2
        }
        row_map = {
            'description of debt obligations': 12, 'amount designated (debt obligations)': 12,
            'amount of original issuance': 12
        }
        
        sort_multiple(data, destination, starting_row, label_row, column_labels, column_map, row_map, current_num)
        
        nonlocal footnote_7, fn7_text, asterisk_check
        if footnote_7:
            footnote = fn7_text
            destination.cell(row=48, column=1, value=footnote)
        if asterisk_check:
            destination.cell(row=45, column=4, value="*")
        return
    
    def section_4():
        destination = current_ATR['Section 4']
        data = Data_Tables['Section 4']
        
        row = 2
        match = 0
        properties = []
        while True:
            num = normalized_value(data.cell(row=row, column=2).value)
            if num is None:
                break
            if num == current_num:
                property_inst = normalized_value(data.cell(row=row, column=3).value)
                properties.append(property_inst)
                match = 1
            row += 1
        
        if match == 1:
            row = 11
            for i in range(len(properties)):
                destination.cell(row=row, column=2).value = properties[i]
                row += 6
        else:
            destination.cell(row=8, column=1).value = 'X'
        return
    
    def section_5():
        class Project:
            def __init__(self, number=None, proj_name=None, p_type=None, ongoing=False, completed=False, est_public=None, 
                         private_investment=None, private_complete=None, public_investment=None, 
                         public_complete=None, project_type=None, new_deal=False, current_payments=None, straddling_fn=None):
                self.num = number
                self.name = proj_name
                self.type = p_type
                self.ong = ongoing
                self.comp = completed
                self.est_pub = est_public
                self.pvt_prev = private_investment
                self.pvt_cmp = private_complete
                self.pub_prev = public_investment
                self.pub_cmp = public_complete
                self.proj_type = project_type
                self.new = new_deal
                self.cur_pmts = current_payments
                self.fn = straddling_fn
        data = Data_Tables['Section 5 pre-22']
        
        column_labels = {
            'tifnum': None, 'prior to 2022': None
        }
        
        old_count = 0
        column_labels = column_match(label_row=1, data=data, column_labels=column_labels)
        tif_row = get_tif_row_single(row=2, num_col=column_labels['tifnum'], data=data, current_num=current_num)
        old_count = data.cell(row=tif_row, column=column_labels['prior to 2022']).value

        
        destination = current_ATR['Section 5']
        data = Data_Tables['Section 5']
        
        column_labels = {
            'project / iga': None, 'type': None, 'tifnum': None, 'project #': None, 
            'rda name normalized': None, 'annual report name': None, 'currentyearnewdeals': None, 
            'ongoing': None, 'complete': None, 'currentyearpmts': None, 'estsubsequentyearpmts': None, 
            'pvt 12-31-99 to yr end': None, 'pvt to completion': None, 'public 11-1-99 to yearend': None,
            'public to completion': None, 'straddling tif footnote #': None
        }
        column_map = {
            'estsubsequentyearpmts': 3, 'pvt 12-31-99 to yr end': 2, 
            'pvt to completion': 4, 'public 11-1-99 to yearend': 2, 'public to completion': 4
        }
        
        # collect data
        column_labels = column_match(label_row=1, data=data, column_labels=column_labels)
        project_list = []
        iga_list = []
        remove_text = False
                
        tif_rows = get_tif_rows(row=3, num_col=column_labels['tifnum'], data=data, current_num=current_num)

        nonlocal footnote_4, footnote_5
        for row in tif_rows:
            number = normalized_value(data.cell(row, column_labels['project #']).value) 
            project_name = normalized_value(data.cell(row, column_labels['annual report name']).value) 
            p_type = normalized_value(data.cell(row, column_labels['type']).value) 
            
            ongoing = False
            if normalized_value(data.cell(row, column_labels['ongoing']).value) not in (0, None):
                ongoing = True
            completed = False
            if normalized_value(data.cell(row, column_labels['complete']).value) not in (0, None):
                completed = True
            new_deal = False
            if normalized_value(data.cell(row, column_labels['currentyearnewdeals']).value) not in (0, None):
                new_deal = True
                
            estimate = normalized_value(data.cell(row, column_labels['estsubsequentyearpmts']).value)
            pub_previous = normalized_value(data.cell(row, column_labels['public 11-1-99 to yearend']).value)
            pub_complete = normalized_value(data.cell(row, column_labels['public to completion']).value) 
            pvt_previous = normalized_value(data.cell(row, column_labels['pvt 12-31-99 to yr end']).value)
            pvt_complete = normalized_value(data.cell(row, column_labels['pvt to completion']).value)
            project_type = normalized_value(data.cell(row, column_labels['project / iga']).value)
            cur_pmts = normalized_value(data.cell(row, column_labels['currentyearpmts']).value)
            straddling_fn = normalized_value(data.cell(row, column_labels['straddling tif footnote #']).value)
            
            if project_type.strip().lower() == 'iga':
                if cur_pmts not in (0, None):
                    iga_list.append(Project(number, project_name, p_type, ongoing, completed, estimate, pvt_previous, pvt_complete, 
                                        pub_previous, pub_complete, project_type, new_deal, cur_pmts, straddling_fn))
            else:
                if project_type.strip().lower() == 'program':
                    footnote_4 = True
                if ongoing:
                    footnote_5 = True
                project_list.append(Project(number, project_name, p_type, ongoing, completed, estimate, pvt_previous, pvt_complete, 
                                        pub_previous, pub_complete, project_type, new_deal, cur_pmts, straddling_fn)) 
        
        try:
            project_list.sort(key=lambda x: x.num)
        except TypeError as e:
            messagebox.showerror(
                "Sorting Error",
                f"Could not sort projects by .num:\n\n{e}\n\n"
                "One of your project numbers values is None.\n\n"
                f"Check Section 5 of the Master Input for TIF {current_num}"
            )
            return
        
        if len(project_list) not in (0, None):
            destination.cell(row=11, column=4, value='X')
            destination.cell(row=12, column=4, value=len(project_list))
            destination.cell(row=13, column=4, value=(len(project_list) - old_count))        
        else:
            destination.cell(row=9, column=4, value='X')
            remove_text = True
        
        # distribute data
        starting_row = 21
        nonlocal footnote_10, footnote_13
        for project in project_list:
            # project name
            destination.merge_cells(start_row=starting_row, start_column=1, end_row=starting_row, end_column=4)
            name = destination.cell(row=starting_row, column=1).value
            name = name + f"  {project.name}"
            if project.proj_type.strip().lower() == 'program':
                name = name + "**"
            if project.ong == True:
                name = name + f" (Project is Ongoing***)"
            elif project.comp == True:
                name = name + f" (Project Completed)"
            if project.fn == 1:
                name = name + " (1)"
            elif project.fn == 3:
                name = name + " (3)"
            
            # update footnote checkers
            if current_num in straddling_tifs:
                if project.fn not in (0, None):
                    if current_num in (53, 95, 107):
                        footnote_13 = True
                    else:
                        footnote_10 = True
            
            if project.fn not in (0, None):    
                value = destination.cell(row=starting_row+2, column=1).value
                if project.fn == 1:
                    value = value + "  (2)"
                elif project.fn == 3:
                    value = value + "  (4)"
                destination.cell(row=starting_row+2, column=1, value=value)
            
            destination.cell(row=starting_row, column=1, value=name)

            # project data
            destination.cell(starting_row+1, column_map['pvt 12-31-99 to yr end']).value = project.pvt_prev
            destination.cell(starting_row+1, column_map['pvt to completion']).value = project.pvt_cmp
            destination.cell(starting_row+2, column_map['public 11-1-99 to yearend']).value = project.pub_prev
            destination.cell(starting_row+2, column_map['estsubsequentyearpmts']).value = project.est_pub
            destination.cell(starting_row+2, column_map['public to completion']).value = project.pub_cmp
            
            starting_row += 5
        
        nonlocal sec5_area
        # remove excess rows
        if len(project_list) < 16: # two pages
            deletion_start = 95
            sec5_area = "A1:D94"
            if len(project_list) < 7: # one page
                deletion_start = 50
                sec5_area = "A1:D49"
                
            destination.delete_rows(deletion_start, amount=(destination.max_row - deletion_start + 1))
        
        
        # Attachment M
        destination = current_ATR['ATTACHMENT M']
        if len(iga_list) == 0:
            del current_ATR['ATTACHMENT M']
            destination = current_ATR['Section 2']
            cell = destination.cell(row=27, column=2, value='X') 
            cell.font = Font(bold=True)
        else:
            if len(iga_list) > 1:
                row_idx = 12
                next_row_idx = row_idx + 1
                row_cells = [destination[row_idx][col] for col in range(len(destination[row_idx]))]  # Get original row cells
                row_height = destination.row_dimensions[row_idx].height  # Save row height
                
                for _ in range(len(iga_list) - 1):
                    destination.insert_rows(next_row_idx)  # Insert a new row before the next row
                    destination.row_dimensions[next_row_idx].height = row_height
                    
                    for col_idx, cell in enumerate(row_cells, start=1):
                        new_cell = destination.cell(row=next_row_idx, column=col_idx, value=cell.value)

                        # Copy formatting
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.protection = copy(cell.protection)
                    next_row_idx += 1
            
            i = 0
            for iga in iga_list:
                destination.cell(row=12+i, column=1).value = iga.name
                cur_type = iga.type.strip().lower()
                if cur_type == "cbe":
                    description = "Improvements to schools"
                    destination.cell(row=12+i, column=2).value = description
                elif cur_type == "cpd":
                    description = "Improvements to parks"
                    destination.cell(row=12+i, column=2).value = description
                elif cur_type == "ccc":
                    description = "Improvements to colleges"
                    destination.cell(row=12+i, column=2).value = description
                elif cur_type == "cta":
                    description = "Improvements to transit system"
                    destination.cell(row=12+i, column=2).value = description
                elif cur_type == "iipd":
                    description = "Improvements to port and terminal facilities"
                    destination.cell(row=12+i, column=2).value = description
                else:
                    description = "N/A"
                    destination.cell(row=12+i, column=2).value = description
                               
                destination.cell(row=12+i, column=3).value = iga.cur_pmts
                destination.cell(row=12+i, column=4).value = None # amount received
                i += 1
                
            destination = current_ATR['Section 2']
            cell = destination.cell(row=27, column=3, value='X') 
            cell.font = Font(bold=True)


        # Footnotes
        destination = current_ATR['Section 5 Footnotes']
        foot_row = 13
        print(remove_text)
        if remove_text:
            for r in range(1, destination.max_row + 1):
                v = destination.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip() == "Project/Program-Specific Notes":
                    destination.delete_rows(r, 1)
                    break
        else:
            if footnote_4:
                footnote = "** Depending on the particular goals of this type of program, the City may: i) make an advance disbursement of the entire public investment amount to the City's program administrator, ii) disburse the amounts through an escrow account, or iii) pay the funds out piecemeal to the program administrator or to the ultimate grantee as each ultimate grantee's work is approved under the program."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            if footnote_5:
                footnote = "*** As of the last date of the reporting fiscal year, the construction of this Project was ongoing; the Private Investment Undertaken and Ratio figures for this Project will be reported on the Annual Report for the fiscal year in which the construction of the Project is completed and the total Private Investment figure is available."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            if footnote_8:
                if current_num in (48, 53, 95):
                    footnote = "(1) This project straddles the Western/Ogden Redevelopment Project Area and the Pilsen Industrial Corridor Redevelopment Project Area and the Midwest Redevelopment Project Area."
                elif current_num in (30, 94):
                    footnote = "(1) This project straddles the Near North Redevelopment Project Area and the Chicago/Kingsbury Redevelopment Project Area."
                elif current_num in (52, 115):
                    footnote = "(1) This project straddles the Kinzie Industrial Corridor Redevelopment Project Area and the Chicago/Central Park Redevelopment Project Area."
                elif current_num in (107, 108):
                    footnote = "(1) This project straddles the Division/Homan Redevelopment Project Area and the Humboldt Redevelopment Project Area."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            if footnote_10:
                footnote = "(2) The Public Investment Undertaken for this project has been funded from increment received from this Area only.  The aggregate amount of Public Investment Undertaken for this Project is the sum of these figures, if any, and the corresponding figures from the other Area or Areas that this Project straddles."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            elif footnote_13:
                footnote = "(2) The Public Investment Undertaken for this project has been funded from increment received from this Area only.  In this case, no increment from this Area was invested in this Project. The aggregate amount of Public Investment Undertaken for this Project is the sum of these figures, if any, and the corresponding figures, if any, from the other Area or Areas that this Project straddles."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            if footnote_9:
                footnote = "(3) This project straddles the Western/Ogden Redevelopment Project Area and the former Near West Redevelopment Project Area."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
            if footnote_11:
                footnote = "(4) This line reports the amounts, if any, that have been or are anticipated to be funded from increment received from this Area only.  The aggregate amount of Public Investment Undertaken for this Project is the sum of these figures, if any, and the corresponding figures from the now-terminated Near West TIF Area."
                destination.cell(row=foot_row, column=1, value=footnote)
                foot_row += 1
        
        return
    
    def section_6():
        destination = current_ATR['Section 6']
        data = Data_Tables['Section 6.2']
        
        column_labels = {
            'tifnum': None, 'projectname': None, 'jobsprojectedtemp': None, 'jobsactualtemp': None, 'jobsprojectedperm': None, 'jobsactualperm': None
        }
        column_map = {
            'projectname': 1,
            'jobsprojectedtemp': 3,
            'jobsactualtemp': 5,
            'jobsprojectedperm': 4,
            'jobsactualperm': 6
        }
        row_map = {
            'projectname': 19,
            'jobsprojectedtemp': 19,
            'jobsactualtemp': 19,
            'jobsprojectedperm': 19,
            'jobsactualperm': 19
        }
        
        label_row = 1
        starting_row = 3
        
        sort_multiple(data, destination, starting_row, label_row, column_labels, column_map, row_map, current_num)

        
        data = Data_Tables['Section 6.3']
        
        column_labels = {
            'tifnum': None, 'projectname': None, 'incrementprojected': None, 'incrementactual': None

        }
        column_map = {
            'projectname': 1,
            'incrementprojected': 3,
            'incrementactual': 5
        }
        row_map = {
            'projectname': 28,
            'incrementprojected': 28,
            'incrementactual': 28
        }
        
        label_row = 1
        starting_row = 3
        
        sort_multiple(data, destination, starting_row, label_row, column_labels, column_map, row_map, current_num)
        
        # destination.cell(row=37, column=1, value="N/A")
        # destination.cell(row=37, column=4, value="N/A")

        return
    
    # # # # # #
    # # # # # 
    # # # #
    # # #
    # #
    #
    
    print("ATR entered")
    
    if getattr(sys, 'frozen', False):
        # Running as bundled exe
        base_dir = os.path.dirname(sys.executable)
    else:
        # Running as plain .py
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    output_dir = os.path.join(base_dir, f"Section Output {reporting_year}")
    os.makedirs(output_dir, exist_ok=True)
    final_dir  = os.path.join(base_dir, f"Final Output {reporting_year}")
    os.makedirs(final_dir, exist_ok=True)
    
    Data_Tables = load_workbook(input_file, data_only=True)
    
    straddling_tifs = [48, 53, 95, 30, 94, 52, 115, 107, 108]
    
    for current_tif in tif_list:
        print("Main loop entered")

        #Footnote global checkers
        footnote_1 = False
        footnote_2 = False
        footnote_3 = False
        footnote_6 = False
        fn6_text = None
        fn7_text = None
        footnote_7 = False
        footnote_4 = False
        footnote_5 = False
        footnote_8 = False
        footnote_9 = False
        footnote_10 = False
        footnote_11 = False
        footnote_13 = False
        # attachment checkers
        # attach_K = False
        # attach_L = False
        
        ijrl = False
        asterisk_check = False
        
        year_suffix = reporting_year[-2:]
        
        problem_sections = ["section 5", "section 5 footnotes", "section 6 footnotes"]
        single_pages = ["section 2", "section 3.1", "section 3.1 other", "section 3.2 B", "section 3.3", "section 6", "section 7", "section 8"]

        current_num = current_tif.num
        current_name = current_tif.name
        try:
            reporting_year_int = int(reporting_year)
        except ValueError:
            print("Invalid reporting year")
        
        if current_num in straddling_tifs:
            footnote_8 = True
            if current_num == 48:
                footnote_9 = True
                footnote_11 = True
        
        data = Data_Tables['Section 1']
        
        name_labels = {
            'tifnum': None, 'filename': None
        }
        name_labels = column_match(1, data, name_labels)
        sec_1_tif_row = get_tif_row_single(2, name_labels['tifnum'], data, current_num)
        atr_name = normalized_value(data.cell(sec_1_tif_row, name_labels['filename']).value)
        
        new_name = f'{atr_name}.xlsm'
        # test_name = f'{atr_name}_TEST.xlsm'
        shutil.copy(template_file, new_name)
        # shutil.copy(template_file, test_name)
        
        current_ATR = load_workbook(new_name, keep_vba=True)
        
        sheets_to_skip = { "section 3.1 other": True }
        
        section_1(section1_list)
        print("Section 1 Passed")
        # current_ATR.save("s1.xlsx")
        section_3_1()
        print("Section 3.1 Passed")
        # current_ATR.save("s3-1.xlsx")
        section_3_2_A()
        print("Section 3.2 A Passed")
        # current_ATR.save("s3-2a.xlsx")
        section_3_2_B()
        print("Section 3.2 B Passed")
        # current_ATR.save("s3-2b.xlsx")
        section_3_3()
        print("Section 3.3 Passed")
        # current_ATR.save("s3-3.xlsx")
        section_4()
        print("Section 4 Passed")
        # current_ATR.save("s4.xlsx")
        sec5_area = "A1:D144"
        section_5()
        print("Section 5 Passed")
        # current_ATR.save("s5.xlsx")
        section_6()
        print("Section 6 Passed")
        # current_ATR.save("s6.xlsx")
        section_2()
        print("Section 2 Passed")
        # current_ATR.save("s2.xlsx")
    
        current_ATR.save(new_name)
        
        form_name = f"{atr_name}AR{year_suffix}"
        bc_docs(output_dir, current_name, form_name, attB_ijrl_file, attC_ijrl_file, 
                attB_tifcorp_file, attC_tifcorp_file, bsigned_file, csigned_file, ijrl)

        # Path to the Excel file you have edited using openpyxl
        excel_file = os.path.abspath(new_name)
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.AutomationSecurity = 1
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.Visible = False
        
        template_wb = excel.Workbooks.Open(template_file)   # open once
        template_wb.Activate()

        excel.Application.Run(f"'{template_wb.Name}'!CopyRichTextIntoFile", excel_file)
        
        template_wb.Close(False)
        wb = excel.Workbooks.Open(excel_file)

        # Iterate through each worksheet in the workbook
        for sheet in wb.Worksheets:
            sname = sheet.Name.lower().strip()
            if sname == "section 3.1 other" and sheets_to_skip['section 3.1 other']:
                continue

            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            
            
            if sname in problem_sections:
                if sname == "section 5":
                    sheet.PageSetup.FitToPagesTall = False
                    sheet.PageSetup.PrintArea = sec5_area
                else:
                    sheet.PageSetup.FitToPagesTall = 1
                    if sname == "section 5 footnotes":
                        sheet.PageSetup.PrintArea = "A1:A19"
                    else:
                        sheet.PageSetup.PrintArea = "A1:B20"
            else:
                sheet.PageSetup.PrintArea = ""
                if sname in single_pages:
                    sheet.PageSetup.FitToPagesTall = 1
                else:
                    sheet.PageSetup.FitToPagesTall = False     # Allow multiple vertical pages                
            
            sheet.PageSetup.CenterVertically = False    
            
            suffix = sheet_suffix(sheet.Name)
            pdf_file = os.path.join(output_dir, f"{atr_name}AR{year_suffix}-{suffix}.pdf")
            
            sheet.Select()
            
            wb.ActiveSheet.ExportAsFixedFormat(0, pdf_file)
            print(f"Exported {atr_name}_{sheet.Name} to {pdf_file}")

        wb.Close(False)
        excel.Application.Quit()
        
        del wb, excel
        import gc, pythoncom
        gc.collect()
        pythoncom.CoUninitialize()             # optional but helps on some setups

        # ---- delete the temporary Excel file --------------------------------
        from pathlib import Path

        try:
            Path(excel_file).unlink()          # or os.remove(excel_file)
            print(f"Deleted {excel_file}")
        except FileNotFoundError:
            pass                               # already gone, fine
        except PermissionError:
            print("File still locked—double-check that Excel really quit.")
    
    merge_by_tif_number(output_dir, year_suffix, delete_sources=True)
            
    if merge:
        written = merge_by_tif_number(output_dir, year_suffix, delete_sources=False, out_suffix=".pdf", all=True)
        for pdf in written:
            shutil.move(str(pdf), os.path.join(final_dir, pdf.name))

