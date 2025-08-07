
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import shutil
import os
import sys
from datetime import date
from Annual_TIF_Report import column_match


# sec 3.1 through 3.1 other prev need to be in tif order!





# Columns to collect from in Data (ReptYear ignored becaused added manually)

section_1     = {'tifnum': None, 'tifname': None, 'approvedate': None, 'expiredate': None, 'filename': None}

section_2     = {'tifnum': None, 'reptyear': None, 'primaryuse': None, 'combomix': None, 'ijrl': None}

section_3_1   = {'tifnum': None, 'reptyear': None, 'taxallocationfundbalance': None, 'proptaxincr-current': None, 
                 'interest-current': None, 'land/bldg-current': None, 'bond-current': None, 'municipal-current': None,
                 'private-current': None, 'totalreceipts': None, 'totalexp/cash': None, 'section3.2atotal': None, 'distributionofsurplus': None,
                 'transfers--municipal': None, 'totalexpend/dist': None, 'netincomecalc': None, 'fundbalancecalc': None}

dest_add_3_1  = {'proptaxincr-current': 5, 'interest-current': 8, 'land/bldg-current': 11, 'bond-current': 14,
                 'municipal-current': 17, 'private-current': 20, 'distributionofsurplus': 24,
                 'transfers--municipal': 25, 'totalexpend/dist': 26}

skip2_list    = {'proptaxincr-current', 'interest-current', 'land/bldg-current', 'bond-current', 'municipal-current'}

skip1_list    = {'taxallocationfundbalance', 'private-current'}

section_3_1_prev = {'proptaxincr-previous': None, 'interest previous': None, 
                    'land building sale previous': None, 'bond proceeds previous': None,
                    'transfers to municipal sources previous': None, 
                    'private sources previous': None}

section_3_1_other = {'tifnum': None, 'reptyear': None, 'noteproceedscurrentyear': None, 'noncompliancepayment': None, 
                     'excessreserverequirement': None, 'babrebate': None, 'collectionreturns': None, 
                     'creditsexpenditures': None}

section_3_1_other_prev = {'prioryearscumulative': None, 'noteproceedscumulative': None, 
                          'noncompliancecum': None, 'excessreservecum': None, 'babrebatecum': None, 
                          'collectionreturnscum': None, 'creditsexpenditurescum': None}

section_3_2_A = {'tifnum': None, 'reptyear': None, 'costofstudies': None, 'administrative cost': None, 
                 'marketing sites': None, 'site preparation costs': None, 'renovation rehab, etc': None, 
                 'public works': None, 'removing contaminants': None, 'jobtraining': None, 
                 'financing costs': None, 'capital costs': None, 'schooldistricts': None, 
                 'librarydistricts': None, 'relocation costs': None, 'inlieu of taxes': None, 
                 'jobtraining-retraining2': None, 'interest cost': None, 'newhousing': None, 
                 'daycare services': None, 'other': None, 'total': None}

section_3_2_B = {"tifnum": None, "reptyear": None, "vendorname": None, "vendorservice": None, "payamt": None} 

section_3_3   = {"tifnum": None, "reptyear": None, "fundbalancecalc": None, "description of debt obligations": None, 
                 "amount of original issuance": None, "amount of additional issuance": None,
                 "amount designated (debt obligations)": None, "descriptions of project costs to be paid": None, 
                 "amount designated _(project costs)": None, "totaldes": None, "surplus/deficit": None}

section_4     = {"reptyear": None, "tifnum": None, "address": None, "property status": None}

section_5     = {'tifnum': None, 'reptyear': None, 'project / iga': None, 'type': None, 'project #': None, 
                 'rda name normalized': None, 'annual report name': None, 'currentyearnewdeals': None, 
                 'ongoing': None, 'complete': None, 'currentyearpmts': None, 'estsubsequentyearpmts': None, 
                 'pvt 12-31-99 to yr end': None, 'pvt to completion': None, 'public 11-1-99 to yearend': None,
                 'public to completion': None}

section_6_2   = {'reptyear': None, 'tifnum': None, 'projectname': None, 'jobsprojectedtemp': None, 'jobsactualtemp': None,
                 'jobsprojectedperm': None, 'jobsactualperm': None}

section_6_3   = {'reptyear': None, 'tifnum': None, 'projectname': None, 'incrementprojected': None, 'incrementactual': None}

attach_A      = {"reptyear": None, "tifnum": None, "tifname": None, "ordinance received": None, "ordinance action": None, "amendment date": None}
attach_D      = {"reptyear": None, "tifnum": None, "projectname": None}
attach_E      = {"reptyear": None, "tifnum": None, "address": None, "propertystatus": None}
attach_H      = {"reptyear": None, "tifnum": None, "tifname": None}

sections = {"Section 1": section_1, "Section 2": section_2, "Section 3.1": section_3_1, "Section 3.1 Previous": section_3_1_prev, "Section 3.1 Other": section_3_1_other,
            "Section 3.1 Other Previous": section_3_1_other_prev, "Section 3.2a": section_3_2_A, "Section 3.2b": section_3_2_B, "Section 3.3": section_3_3, 
            "Section 4": section_4, "Section 5": section_5, "Section 6.2": section_6_2, "Section 6.3": section_6_3, "Attachment A": attach_A, "Attachment D": attach_D, 
            "Attachment E": attach_E, "Attachment H": attach_H} # cap mismatch on 3.2a and 3.2b

calculated = {"Section 3.1", "Section 3.1 Previous", "Section 3.1 Other", "Section 3.1 Other Previous"}

start_rows = {"Section 1": 2, "Section 2": 2, "Section 3.1": 4, "Section 3.1 Previous": 3, "Section 3.1 Other": 4,
              "Section 3.1 Other Previous": 2, "Section 3.2a": 6, "Section 3.2b": 3, "Section 3.3": 4, 
              "Section 4": 2, "Section 5": 3, "Section 6.2": 3, "Section 6.3": 3, "Attachment A": 2, "Attachment D": 2, 
              "Attachment E": 2, "Attachment H": 2}


# def column_match(label_row: int, data):
#     """Selects the "tifnum" column in the excel file

#     Args:
#         label_row (int): row that holds the labels of the columns (in data)
#         data: sheet that contains the input data

#     Returns:
#         col: "tifnum" column
#     """
#     for cell in data[label_row]:
#         if cell.value:
#             label = str(cell.value).strip().lower()
#             if label == "tifnum":
#                 col = cell.column
#     return col

def set_data_length(data, tif_col: int, row_start: int):
    """Determines the how far down the column to go for the data

    Args:
        data (path): sheet that contains the input data
        tif_col (int): column that holds the TIF numbers
        row_start (int): first row that holds the data we're looking for in the sheet

    Returns:
        length (int): vertical length of data in the sheet
    """
    length = 0
    row = row_start
    while True:
        cell = data.cell(row=row, column=tif_col)
        if cell.value != None:
            length += 1
            row    += 1
        else:
            break
        
    return length

def fill_date(destination, row: int, col: int, reporting_year: str, length: int):
    while length > 0:
        destination.cell(row, col, value=reporting_year)
        row    += 1
        length -= 1
    
    return

def get_column_data(data, col: int, start_row: int, length: int):
    """Collect the data from the input sheet

    Args:
        data (path): sheet that contains the input data
        col (int): lettered column  in sheet where data is being collected like "A" or "B" (must be single column)
        label_row (int): row that holds the labels of the columns (in data)
        start_row (int): first row that holds the data we're looking for in the sheet
        length (int): vertical length of data in the sheet

    Returns:
        values (list): list of values in a column of data
    """
    # cells  = data[col][start_row-1 : start_row+length]
    # values = [cell.value for cell in cells]
    
    # return values
    return [
        data.cell(row=r, column=col).value
        for r in range(start_row, start_row + length)
    ]

def fill_column(destination, col, start_row, values):
    for value in values:
        destination.cell(start_row, col, value=value)
        start_row += 1
    return

def copy_columns(src_ws, dst_ws, mapping, src_start, dst_start_row, reporting_year):
    """
    Args:
        src_ws        : openpyxl worksheet you're reading from
        dst_ws        : openpyxl worksheet you're writing to
        mapping_dict  : { logical_name: src_col_index, … }
        src_start     : the first row in src_ws to read
        dst_start_row : the first row in dst_ws to write
    """
    length = set_data_length(src_ws, mapping["tifnum"], src_start)
    if length == 0:
        return

    # 2) iterate *in the order of your mapping keys*
    for dst_col, field in enumerate(mapping.keys(), start=1):
        src_col = mapping[field]
        if field == "reptyear":
            fill_date(dst_ws, dst_start_row, dst_col, reporting_year, length)
        elif field == "total":
            continue
        elif src_col is not None:
            # grab the entire column from src and write it downwards in dst
            for i in range(length):
                val = src_ws.cell(row=src_start + i, column=src_col).value
                dst_ws.cell(row=dst_start_row + i, column=dst_col, value=val)

def _num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    try:
        # strip commas, dollar‐signs, whatever—and float it
        return float(str(v).replace(',', '').replace('$',''))
    except ValueError:
        return 0
    

def populate_sheet(master_input, final_table, reporting_year):
    """Populates an entire column of data from one workbook to another

    Args:
        master_input (path): workbook that contains the input data
        destination (path): sheet that contains the output data
    """
    for sec_name, section in sections.items():
        length = 0
        if sec_name in calculated:
            if sec_name == "Section 3.1":
                print("entered 3.1")
                prev_name = "Section 3.1 Previous"
                data = master_input[sec_name]
                prev_data = master_input[prev_name]
                destination = final_table[sec_name]

                matched_section = column_match(1, data, sections[sec_name].copy())
                prev_matched_section = column_match(1, prev_data, sections["Section 3.1 Previous"].copy())
                length = set_data_length(data, matched_section["tifnum"], start_rows[sec_name])
                
                i = 1
                for col_name, col_num in matched_section.items():
                    if col_name == "reptyear":
                        fill_date(destination, 2, i, reporting_year, length)
                    elif col_name == "section3.2atotal" or col_name == "totalexpend/dist":
                        if col_name == "totalexpend/dist":
                            i += 1
                        continue
                    elif length != 0:
                        if col_name == "totalexp/cash":
                            data = master_input["Section 3.2a"]

                            labels = column_match(1, data, sections['Section 3.2a'])
                            values = get_column_data(data, labels['total'], start_rows["Section 3.2a"], length)
                            fill_column(destination, i, 2, values)
                            
                            data = master_input[sec_name]
                        else:
                            values = get_column_data(data, col_num, start_rows[sec_name], length)
                            fill_column(destination, i, 2, values)
                        
                    if col_name in skip2_list:
                        i += 3
                    elif col_name in skip1_list:
                        i += 2
                    else:
                        i += 1

                i = 4
                for col_name, col_num in prev_matched_section.items():
                    if length != 0:
                        values = get_column_data(prev_data, col_num, start_rows[prev_name], length)
                        fill_column(destination, i, 2, values)
                    i += 3

                i = 0
                values = get_column_data(data, matched_section['section3.2atotal'], start_rows[sec_name], length)
                for value in values:
                    value1 = _num(destination.cell(row=i+2, column=(dest_add_3_1['proptaxincr-current'])-1).value) + _num(destination.cell(row=i+2, column=dest_add_3_1['proptaxincr-current']).value)
                    value2 = _num(destination.cell(row=i+2, column=(dest_add_3_1['interest-current'])-1).value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['interest-current'])).value)
                    value3 = _num(destination.cell(row=i+2, column=(dest_add_3_1['land/bldg-current'])-1).value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['land/bldg-current'])).value)
                    value4 = _num(destination.cell(row=i+2, column=(dest_add_3_1['bond-current'])-1).value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['bond-current'])).value)
                    value5 = _num(destination.cell(row=i+2, column=(dest_add_3_1['municipal-current'])-1).value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['municipal-current'])).value)
                    value6 = _num(destination.cell(row=i+2, column=(dest_add_3_1['private-current'])-1).value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['private-current'])).value)
                    destination.cell(row=i+2, column=(dest_add_3_1['proptaxincr-current'])+1, value=value1)
                    destination.cell(row=i+2, column=(dest_add_3_1['interest-current'])+1, value=value2)
                    destination.cell(row=i+2, column=(dest_add_3_1['land/bldg-current'])+1, value=value3)
                    destination.cell(row=i+2, column=(dest_add_3_1['bond-current'])+1, value=value4)
                    destination.cell(row=i+2, column=(dest_add_3_1['municipal-current'])+1, value=value5)
                    destination.cell(row=i+2, column=(dest_add_3_1['private-current'])+1, value=value6)
                    # destination.cell(row=i+2, column=23).value = value1 + value2 + value3 + value4 + value5 + value6
                    
                    value7 = _num(value) + _num(destination.cell(row=i+2, column=(dest_add_3_1['distributionofsurplus'])).value) + _num(destination.cell(row=i+2, column=dest_add_3_1['transfers--municipal']).value)
                    destination.cell(row=i+2, column=(dest_add_3_1['totalexpend/dist']), value=value7)

                    i += 1
                

            elif sec_name == "Section 3.1 Other":
                print("entered 3.1 Other")
                prev_name = "Section 3.1 Other Previous"
                data = master_input[sec_name]
                prev_data = master_input[prev_name]
                destination = final_table[sec_name]
                
                matched_section = column_match(1, data, sections[sec_name].copy())
                prev_matched_section = column_match(1, prev_data, sections[prev_name].copy())
                length = set_data_length(data, matched_section["tifnum"], start_rows[sec_name])
                
                i = 1
                for col_name, col_num in matched_section.items():
                    if col_name == "reptyear":
                        fill_date(destination, 2, i, reporting_year, length)
                    elif length != 0:
                        values = get_column_data(data, col_num, start_rows[sec_name], length)
                        fill_column(destination, i, 2, values)
                    
                    if col_name == "tifnum":
                        i += 1
                    else:
                        i += 2

                i = 3
                for col_name, col_num in prev_matched_section.items():
                    if length != 0:
                        values = get_column_data(prev_data, col_num, start_rows[prev_name], length)
                        j = 0
                        for value in values:
                            destination.cell(j+2, i).value = _num(value) + _num(destination.cell(j+2, i-1).value)
                            j += 1
                    i += 2


        else:
            # data = master_input[sec_name]
            # destination = final_table[sec_name]

            # matched_section = column_match(1, data, section)
            # length = set_data_length(data, matched_section["tifnum"], start_rows[sec_name])
            # i = 1

            # for col_name, col_num in matched_section.items():
            #     if col_name == "reptyear":
            #         fill_date(destination, 2, i, reporting_year, length)
            #     elif length != 0:
            #         values = get_column_data(data, col_num, start_rows[sec_name], length)
            #         fill_column(destination, i, 2, values)

            #     i += 1
            print(f"entered {sec_name}")
            data        = master_input[sec_name]
            destination = final_table[sec_name]
            mapping     = sections[sec_name].copy()
            mapping     = column_match(1, data, mapping)
            src_row     = start_rows[sec_name]
            dst_row     = 2  # or wherever your output always begins

            copy_columns(data, destination, mapping, src_row, dst_row, reporting_year)
    return
    

def Data_Tables(reporting_year_input, input_file, template_file):
    print("Data Tables entered")

    reporting_year = _num(reporting_year_input)
    today = date.today()
    date_str = today.strftime("%m.%d.%y") 
    master_input = load_workbook(input_file, data_only=True)
    new_name = f"Data Tables {date_str}.xlsx"
    shutil.copy(template_file, new_name)  # CHANGE IF FILE CHANGES
    final_table = load_workbook(new_name)

    populate_sheet(master_input, final_table, reporting_year)

    final_table.save(new_name)


